from fastapi import APIRouter, Request, Form, HTTPException, UploadFile, File
from app.config import logger
from app.dependencies import user_sessions, get_user_client
from app.utils.transliteration import transliterate
from app.utils.validation import is_valid_email
from app.utils.excel import parse_excel_row, parse_fio, parse_groups
from app.services.yopass import create_yopass_link
from app.models.user import UserCreate
from typing import Optional, Dict, Any
import openpyxl
from io import BytesIO


router = APIRouter()


@router.get("/api/v1/users/{username}")
def get_user(username: str, request: Request) -> Dict[str, Any]:
    """
    Получение информации о пользователе
    Аналог в FreeIPA: ipa user-show --all username
    """

    try:
        # Получаем клиент из сессии
        client = get_user_client(request)
        
        # Получаем информацию о пользователе
        user = client._request("user_show", args=[username], params={"all": True})
        return user
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка получения пользователя: {str(e)}"
        )

@router.post("/api/v1/users/{username}/delete")
def delete_user(username: str, request: Request) -> Dict[str, str]:
    """
    Обычное удаление пользователя в FreeIPA

    Удаляет одного пользователя

    Использовать с умом т.к безвозвратно удаляет пользователя в FreeIPA
    """
    try:
        session_id = request.cookies.get("ipa_session")
        admin = user_sessions.get(session_id, {}).get("username", "unknown")
        logger.warning(f"USER_DELETE: {username} by {admin}")

        client = get_user_client(request)
        result = client._request("user_del", args=[username], params={})
        
        logger.info(f"USER_DELETE SUCCESS: {username}")
        return {
            "username": username,
            "message": f"Пользователь {username} успешно удалён",
            "status": "deleted"
        }

    except Exception as e:
        logger.error(f"USER_DELETE FAILED: {username} - {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка удаления пользователя: {str(e)}"
        )

@router.post("/api/v1/users/{username}/disable")
def disable_user(username: str, request: Request) -> Dict[str, str]:
    """
    Обычное выключение пользователя в FreeIPA

    Выключение одного пользователя
    """
    try:
        client = get_user_client(request)
        result = client._request("user_disable", args=[username], params={})

        return {
            "username": username,
            "message": f"Пользователь {username} успешно отключен",
            "status": "disable"
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка отключения пользователя: {str(e)}"
        )

@router.post("/api/v1/users/{username}/enable")
def enable_user(username: str, request: Request) -> Dict[str, str]:
    """
    Активация юзера в FreeIPA

    Активация одного пользователя
    """
    try:
        client = get_user_client(request)
        result = client._request("user_enable", args=[username], params={})

        return {
            "username": username,
            "message": f"Пользователь {username} успешно включен",
            "status": "enable"
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка активации пользователя: {str(e)}"
        )

@router.post("/api/v1/users/{username}/reset-password")
def reset_password(username: str, request: Request) -> Dict[str, Any]:
    """
    Сброс пароля пользователя в FreeIPA

    Пароль генерируется автоматически и возвращается далее для работы

    Также возвращается дата до какого числа будет действовать пароль `"expiration": result['result'].get('mail', [None])[0]`

    Вводим только логин УЗ
    """
    try:
        session_id = request.cookies.get("ipa_session")
        admin = user_sessions.get(session_id, {}).get("username", "unknown")
        logger.warning(f"PASSWORD_RESET: {username} by {admin}")

        client = get_user_client(request)
        result = client._request("user_mod", args=[username], params={"random": True})

        password = result['result']['randompassword']

        yopass_link = create_yopass_link(username,password)

        response = {
            "username": username,
            "password": password,
            "yopass_link": yopass_link,
            "expiration": result['result'].get('krbpasswordexpiration', [None])[0],
            "message": f"Пароль пользователя {username} успешно сброшен"
        }

        logger.info(f"PASSWORD_RESET SUCCESS: {username}")
        return response

    except Exception as e:
        logger.error(f"PASSWORD_RESET FAILED: {username} - {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка сброса пароля: {str(e)}"
        )



@router.post("/api/v1/creat-users")
def create_user(user: UserCreate, request: Request) -> Dict[str, Any]:
    """
    Создает пользователя c передачей цельного JSON файла

    Можно указать список групп для автоматического добавления:

    {
        "first_name": "Иван",
        "last_name": "Иванов",
        "email": "ivan@test.com",
        "groups": ["admins", "developers"]
    }

    ВАЖНО: Если указаны группы и ни одна не добавилась - пользователь будет удален и вернется ошибка.
    """
    try:
        # Транслитерация для поддержки кириллицы
        first_name_en = transliterate(user.first_name).lower()
        last_name_en = transliterate(user.last_name).lower()
        username = f"{first_name_en}.{last_name_en}"
        full_name = f"{user.first_name} {user.last_name}"

        session_id = request.cookies.get("ipa_session")
        admin = user_sessions.get(session_id, {}).get("username", "unknown")
        logger.info(f"USER_CREATE: {username} ({user.email}) by {admin}")

        client = get_user_client(request)

        # Создаём пользователя тут ipa user_add
        result = client._request(
            "user_add",
            args=[username],
            params={
                "givenname": user.first_name,
                "sn": user.last_name,
                "cn": full_name,
                "mail": user.email,
                "title": user.title,
                "telephonenumber": user.phone,
                "random": True,
            }
        )

        password = result['result']['randompassword']

        # Тут вызываю новый метод group_add_member т.к в user_add нет такого функционала
        added_groups = []
        failed_groups = []

        for group in user.groups:
            try:
                client._request(
                    "group_add_member",
                    args=[group],
                    params={"user": username}
                )
                added_groups.append(group)
            except Exception as e:
                failed_groups.append({"group": group, "error": str(e)})

        response = {
            "username": username,
            "password": password,
            "email": user.email, # Сразу добавил чтобы выводил почту чтобы передавать её дальше для генерации ссылки в passoc и отправки
            "message": f"Пользователь {username} создан"
        }

        # Проверяем результат добавления в группы
        if user.groups:
            # Если группы были указаны, но НИ ОДНА не добавилась - откатываю создание пользователя
            if len(added_groups) == 0:
                try:
                    client._request("user_del", args=[username], params={})
                except Exception as e:
                    logger.warning(f"Failed to delete user {username} during rollback: {e}")

                raise HTTPException(
                    status_code=500,
                    detail=f"Пользователь создан, но не удалось добавить ни в одну группу. Пользователь удален. Ошибки: {failed_groups}"
                )

            # Если хотя бы одна группа добавилась - возвращаем успех с информацией
            response["groups"] = {
                "added": added_groups,
                "failed": failed_groups
            }

        logger.info(f"USER_CREATE SUCCESS: {username} with groups {added_groups}")
        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"USER_CREATE FAILED: {username} - {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/api/v1/users/create-form")
def create_user_form(
    request: Request,
    first_name: str = Form(...),
    last_name: str = Form(...),
    email: str = Form(...),
    title: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    groups: Optional[str] = Form(None)
):
    """
    Это аналог ручки /api/v1/creat-users только уже не в формате JSON чтобы ручками его не создавать
    

    ВАЖНО: Если указаны группы и ни одна не добавилась - пользователь будет удален и вернется ошибка.
    """
    try:
        # Транслитерация для поддержки кириллицы
        first_name_en = transliterate(first_name).lower()
        last_name_en = transliterate(last_name).lower()
        username = f"{first_name_en}.{last_name_en}"
        full_name = f"{first_name} {last_name}"

        client = get_user_client(request)

        result = client._request(
            "user_add",
            args=[username],
            params={
                "givenname": first_name,
                "sn": last_name,
                "cn": full_name,
                "mail": email,
                "title": title,
                "telephonenumber": phone,
                "random": True,
            }
        )

        password = result['result']['randompassword']

        added_groups = []
        failed_groups = []

        # Парсим строку с группами (разделенные запятыми)
        groups_list = []
        if groups and groups.strip(): # Проверяю что groups не null
            groups_list = [g.strip() for g in groups.split(',') if g.strip()]   # тут split убирает пробелы т.к 100 процентов будут ошибки и split чтобы разбивать если несколько групп то есть ["admins", "dev", "ops"]

        for group in groups_list:
            try:
                client._request(
                    "group_add_member",
                    args=[group],
                    params={"user": username}
                )
                added_groups.append(group)
            except Exception as e:
                failed_groups.append({"group": group, "error": str(e)})

        response = {
            "username": username,
            "password": password,
            "email": email,
            "message": f"Пользователь {username} создан"
        }

        if groups_list:
            if len(added_groups) == 0:
                try:
                    client._request("user_del", args=[username], params={})
                except Exception as e:
                    logger.warning(f"Failed to delete user {username} during rollback: {e}")
                raise HTTPException(
                    status_code=500,
                    detail=f"Пользователь создан, но не удалось добавить ни в одну группу. Пользователь удален. Ошибки: {failed_groups}"
                )
            response["groups"] = {
                "added": added_groups,
                "failed": failed_groups
            }

        return response

    except HTTPException:
        raise  # перебрасываем HTTPException дальше 
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
@router.get("/api/v1/users/search-by-email/{email}")
def search_user_by_email(email: str, request: Request) -> Dict[str, Any]:
    """
    Поиск пользователя по email (для отладки)

    Показывает что именно возвращает FreeIPA при поиске
    """
    try:
        client = get_user_client(request)

        # Пробуем разные варианты поиска
        results = {}

        # 1. Точное совпадение
        try:
            search_exact = client._request(
                "user_find",
                args=[],
                params={"mail": email}
            )
            results["exact_match"] = {
                "query": email,
                "type": str(type(search_exact)),
                "is_dict": isinstance(search_exact, dict),
                "is_list": isinstance(search_exact, list),
                "data": search_exact
            }
        except Exception as e:
            results["exact_match"] = {"error": str(e)}

        # 2. Lowercase
        try:
            search_lower = client._request(
                "user_find",
                args=[],
                params={"mail": email.lower()}
            )
            results["lowercase"] = {
                "query": email.lower(),
                "type": str(type(search_lower)),
                "is_dict": isinstance(search_lower, dict),
                "is_list": isinstance(search_lower, list),
                "data": search_lower
            }
        except Exception as e:
            results["lowercase"] = {"error": str(e)}

        # 3. Поиск по username (если это вдруг username)
        try:
            user_show = client._request("user_show", args=[email.split('@')[0]], params={})
            results["by_username"] = user_show.get('result', {})
        except Exception as e:
            results["by_username"] = {"error": str(e)}

        return results

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/v1/users/validate-excel")
async def validate_excel(request: Request, file: UploadFile = File(...)):
    """
    Валидация Excel файла перед массовым созданием пользователей

    Проверяет:
    - Формат email
    - Конфликты username (уже существует в FreeIPA)
    - Конфликты email (уже существует в FreeIPA)
    - Дубликаты email внутри файла
    - Существование групп
    - Корректность ФИО (минимум 2 слова)

    Возвращает детальный отчёт БЕЗ создания пользователей
    """
    try:
        # Проверяем авторизацию
        client = get_user_client(request)

        session_id = request.cookies.get("ipa_session")
        admin = user_sessions.get(session_id, {}).get("username", "unknown")
        logger.info(f"VALIDATE_EXCEL: Started by {admin}")

        # Читаем Excel файл
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents))
        sheet = workbook.active

        conflicts = []
        warnings = []
        would_create = 0
        emails_in_file = {}  # Для отслеживания дубликатов внутри файла

        # Получаем список всех существующих пользователей из FreeIPA
        existing_users = client._request("user_find", args=[], params={"all": True})
        existing_usernames = {u['uid'][0] for u in existing_users['result']}

        # Собираем существующие email (безопасно)
        existing_emails = set()
        for u in existing_users['result']:
            mail = u.get('mail')
            if mail and isinstance(mail, list) and len(mail) > 0 and mail[0]:
                existing_emails.add(mail[0].lower())

        # Кэш для проверки групп (чтобы не проверять одну группу несколько раз)
        groups_cache = {}

        # Проходим по строкам
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Пропускаем пустые строки
            if not row or not row[0]:
                continue

            try:
                # Извлекаем данные
                fio = str(row[0]).strip() if row[0] else ""
                email = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                phone = str(row[2]).strip() if len(row) > 2 and row[2] else None
                title = str(row[3]).strip() if len(row) > 3 and row[3] else None
                groups_str = str(row[4]).strip() if len(row) > 4 and row[4] else ""

                # Проверка 1: ФИО заполнено
                if not fio:
                    conflicts.append({
                        "row": row_num,
                        "error": "ФИО не заполнено"
                    })
                    continue

                # Проверка 2: Email заполнен
                if not email:
                    conflicts.append({
                        "row": row_num,
                        "fio": fio,
                        "error": "Email не заполнен"
                    })
                    continue

                # Проверка 3: Email валидный
                if not is_valid_email(email):
                    conflicts.append({
                        "row": row_num,
                        "fio": fio,
                        "error": f"Невалидный email: {email}"
                    })
                    continue

                # Проверка 4: Email дубликат внутри файла
                email_lower = email.lower()
                if email_lower in emails_in_file:
                    conflicts.append({
                        "row": row_num,
                        "fio": fio,
                        "error": f"Дубликат email {email} (уже в строке {emails_in_file[email_lower]})"
                    })
                    continue
                emails_in_file[email_lower] = row_num

                # Парсим ФИО
                fio_parts = fio.split()
                if len(fio_parts) < 2:
                    conflicts.append({
                        "row": row_num,
                        "fio": fio,
                        "error": "ФИО должно содержать минимум Фамилию и Имя"
                    })
                    continue

                last_name = fio_parts[0]
                first_name = fio_parts[1]

                # Генерируем username
                last_name_en = transliterate(last_name).lower()
                first_name_en = transliterate(first_name).lower()
                username = f"{first_name_en}.{last_name_en}"

                # Проверка 5 и 6: Собираем все конфликты для этой строки
                row_errors = []

                # Проверка 5: Username уже существует в FreeIPA
                if username in existing_usernames:
                    row_errors.append(f"Username '{username}' уже существует в FreeIPA")

                # Проверка 6: Email уже существует в FreeIPA
                if email_lower in existing_emails:
                    row_errors.append(f"Email '{email}' уже существует в FreeIPA")

                # Если есть конфликты - добавляем и пропускаем строку
                if row_errors:
                    conflicts.append({
                        "row": row_num,
                        "fio": fio,
                        "username": username,
                        "email": email,
                        "error": "; ".join(row_errors)
                    })
                    continue

                # Проверка 7: Существование групп
                if groups_str:
                    groups_list = [g.strip() for g in groups_str.split(',') if g.strip()]
                    non_existing_groups = []

                    for group in groups_list:
                        # Проверяем кэш
                        if group not in groups_cache:
                            try:
                                client._request("group_show", args=[group])
                                groups_cache[group] = True
                            except Exception:
                                groups_cache[group] = False

                        if not groups_cache[group]:
                            non_existing_groups.append(group)

                    if non_existing_groups:
                        conflicts.append({
                            "row": row_num,
                            "fio": fio,
                            "username": username,
                            "error": f"Группы не существуют: {', '.join(non_existing_groups)}"
                        })
                        continue

                # Предупреждения (не блокируют создание)
                if not phone:
                    warnings.append({
                        "row": row_num,
                        "fio": fio,
                        "username": username,
                        "message": "Телефон не заполнен"
                    })

                if not title:
                    warnings.append({
                        "row": row_num,
                        "fio": fio,
                        "username": username,
                        "message": "Должность не заполнена"
                    })

                # Если всё ок - считаем как валидного
                would_create += 1

            except Exception as e:
                conflicts.append({
                    "row": row_num,
                    "fio": fio if 'fio' in locals() else "unknown",
                    "error": f"Неожиданная ошибка: {str(e)}"
                })

        # Формируем результат
        total_rows = sheet.max_row - 1  # Минус заголовок
        valid = len(conflicts) == 0

        result = {
            "valid": valid,
            "total_rows": total_rows,
            "would_create": would_create,
            "conflicts_count": len(conflicts),
            "warnings_count": len(warnings),
            "conflicts": conflicts,
            "warnings": warnings
        }

        logger.info(f"VALIDATE_EXCEL: Completed by {admin} - Valid: {valid}, Would create: {would_create}, Conflicts: {len(conflicts)}")
        return result

    except Exception as e:
        logger.error(f"VALIDATE_EXCEL: Critical error - {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка валидации Excel файла: {str(e)}"
        )

@router.post("/api/v1/users/bulk-create-from-excel")
async def bulk_create_from_excel(request: Request, file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Парсинг excel и создание пользователя
    """
    try:
        # Сначала проверяем авторизацию (до чтения файла!)
        client = get_user_client(request)

        session_id = request.cookies.get("ipa_session")
        admin = user_sessions.get(session_id, {}).get("username", "unknown")
        logger.info(f"BULK_CREATE_EXCEL: Started by {admin}")

        # Читаем Excel файл (только если авторизован)
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents)) # превращаем биты в читаемый файл
        sheet = workbook.active # Активный листы

        # Проверяем доступность Yopass ДО начала создания пользователей
        try:
            test_link = create_yopass_link("test", "test123")
            logger.info(f"BULK_CREATE_EXCEL: Yopass check OK - {test_link}")
        except Exception as e:
            logger.error(f"BULK_CREATE_EXCEL: Yopass unavailable - {str(e)}")
            raise HTTPException(
                status_code=503,
                detail=f"Yopass недоступен: {str(e)}. Создание пользователей отменено."
            )

        results = {"success": [], "failed": []}

        # Проходим по строкам (пропускаем первую - заголовки)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Пропускаем пустые строки
            if not row or not row[0]:
                continue

            try:
                # Парсим строку Excel
                data = parse_excel_row(row)
                fio = data["fio"]
                email = data["email"]

                # Валидация обязательных полей
                if not fio:
                    results["failed"].append({"row": row_num, "error": "ФИО не заполнено"})
                    continue

                if not email:
                    results["failed"].append({"row": row_num, "fio": fio, "error": "Email не заполнен"})
                    continue

                if not is_valid_email(email):
                    results["failed"].append({"row": row_num, "fio": fio, "error": f"Невалидный email: {email}"})
                    continue

                # Парсим ФИО
                fio_parsed = parse_fio(fio)
                if not fio_parsed:
                    results["failed"].append({"row": row_num, "fio": fio, "error": "ФИО должно содержать минимум Фамилию и Имя"})
                    continue

                last_name, first_name, username = fio_parsed

                # Собираем все ошибки валидации для этой строки
                row_errors = []

                # Проверка: Username уже существует в FreeIPA
                try:
                    client._request("user_show", args=[username])
                    # Если не упало - значит пользователь существует
                    row_errors.append(f"Username '{username}' уже существует в FreeIPA")
                except Exception:
                    # Пользователь не найден - можно создавать
                    pass

                # Проверка: Email уже существует в FreeIPA
                try:
                    # Ищем пользователей с таким email
                    email_check = client._request("user_find", args=[], params={"mail": email})
                    if email_check['result']:
                        # Найден пользователь с таким email
                        existing_username = email_check['result'][0]['uid'][0]
                        row_errors.append(f"Email '{email}' уже используется пользователем {existing_username}")
                except Exception:
                    # Ошибка поиска - игнорируем и продолжаем
                    pass

                # Парсим группы
                groups_list = parse_groups(data["groups_str"])

                # Проверка: Существование всех групп
                if groups_list:
                    non_existing_groups = []
                    for group in groups_list:
                        try:
                            client._request("group_show", args=[group])
                        except Exception:
                            non_existing_groups.append(group)

                    if non_existing_groups:
                        row_errors.append(f"Группы не существуют: {', '.join(non_existing_groups)}")

                # Если есть любые ошибки валидации - не создаём пользователя
                if row_errors:
                    results["failed"].append({
                        "row": row_num,
                        "fio": fio,
                        "username": username,
                        "email": email,
                        "error": "; ".join(row_errors)
                    })
                    continue

                # Создаём пользователя в FreeIPA
                result = client._request(
                    "user_add",
                    args=[username],
                    params={
                        "givenname": first_name,
                        "sn": last_name,
                        "cn": fio,
                        "mail": email,
                        "title": data["title"],
                        "telephonenumber": data["phone"],
                        "random": True,
                    }
                )

                password = result['result']['randompassword']

                yopass_link = create_yopass_link(username, password)

                # Добавляем в группы
                added_groups = []
                failed_groups = []

                for group in groups_list:
                    try:
                        client._request(
                            "group_add_member",
                            args=[group],
                            params={"user": username}
                        )
                        added_groups.append(group)
                    except Exception as e:
                        failed_groups.append({"group": group, "error": str(e)})

                # Если валидация прошла успесното добавляем сюда
                success_entry = {
                    "row": row_num,
                    "fio": fio,
                    "username": username,
                    "email": email,
                    "password": password,
                    "yopass_link": yopass_link
                }

                if groups_list:
                    success_entry["groups"] = {
                        "added": added_groups,
                        "failed": failed_groups
                    }

                results["success"].append(success_entry)
                logger.info(f"BULK_CREATE_EXCEL: Created {username} from row {row_num}")

            except Exception as e:
                results["failed"].append({
                    "row": row_num,
                    "fio": fio if 'fio' in locals() else "unknown",
                    "error": str(e)
                })
                logger.error(f"BULK_CREATE_EXCEL: Failed row {row_num} - {str(e)}")

        logger.info(f"BULK_CREATE_EXCEL: Completed by {admin} - Success: {len(results['success'])}, Failed: {len(results['failed'])}")

        return results

    except Exception as e:
        logger.error(f"BULK_CREATE_EXCEL: Critical error - {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка обработки Excel файла: {str(e)}"
        )
