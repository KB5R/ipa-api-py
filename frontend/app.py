import streamlit as st
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from typing import Optional, List
from dotenv import load_dotenv
import os

load_dotenv()
API_URL = os.getenv("API_URL")

for key, default in [
    ("logged_in", False),
    ("session_cookie", None),
    ("username", None),
    ("pwd_preview", None),
    ("pwd_results", None),
    ("disable_preview", None),
    ("disable_results", None),
    ("state_preview", None),
    ("state_results", None),
    ("state_mode_used", None),
    ("state_mode_used_exec", None),
    ("csv_data", None),
    ("groups_list", None),
    ("search_results", None),
    ("search_action_result", None),
    ("show_passwords_mode", False),
]:
    if key not in st.session_state:
        st.session_state[key] = default


# === API FUNCTIONS ===

def login(username: str, password: str) -> bool:
    try:
        response = requests.post(
            f"{API_URL}/api/v1/session/login",
            data={"username": username, "password": password}
        )
        if response.status_code == 403:
            st.error("Доступ запрещён: требуется группа helpdesk или admins")
            return False
        if response.ok:
            st.session_state.session_cookie = response.cookies.get("ipa_session")
            st.session_state.logged_in = True
            st.session_state.username = username
            return True
        st.error(f"Ошибка входа: {response.json().get('detail', 'Неизвестная ошибка')}")
        return False
    except Exception as e:
        st.error(f"Ошибка подключения: {e}")
        return False


def logout():
    try:
        if st.session_state.session_cookie:
            requests.post(
                f"{API_URL}/api/v1/session/logout",
                cookies={"ipa_session": st.session_state.session_cookie}
            )
    finally:
        st.session_state.logged_in = False
        st.session_state.session_cookie = None
        st.session_state.username = None


def get_cookies() -> dict:
    return {"ipa_session": st.session_state.session_cookie}


def bulk_reset_with_yopass(identifiers: List[str]) -> Optional[dict]:
    try:
        response = requests.post(
            f"{API_URL}/api/v1/users/bulk-reset-password-with-yopass",
            json=identifiers,
            cookies=get_cookies()
        )
        if response.ok:
            return response.json()
        st.error(f"Ошибка: {response.json().get('detail', 'Неизвестная ошибка')}")
        return None
    except Exception as e:
        st.error(f"Ошибка: {e}")
        return None


def bulk_reset_plain(identifiers: List[str]) -> Optional[dict]:
    try:
        response = requests.post(
            f"{API_URL}/api/v1/users/bulk-reset-password",
            json=identifiers,
            cookies=get_cookies()
        )
        if response.ok:
            return response.json()
        st.error(f"Ошибка: {response.json().get('detail', 'Неизвестная ошибка')}")
        return None
    except Exception as e:
        st.error(f"Ошибка: {e}")
        return None


def bulk_reset_from_excel(file) -> Optional[dict]:
    try:
        response = requests.post(
            f"{API_URL}/api/v1/users/bulk-reset-password-from-excel",
            files={"file": file},
            cookies=get_cookies()
        )
        if response.ok:
            return response.json()
        st.error(f"Ошибка: {response.json().get('detail', 'Неизвестная ошибка')}")
        return None
    except Exception as e:
        st.error(f"Ошибка: {e}")
        return None


def bulk_disable_preview_list(identifiers: List[str]) -> Optional[dict]:
    try:
        response = requests.post(
            f"{API_URL}/api/v1/users/bulk-disable-preview-list",
            json=identifiers,
            cookies=get_cookies()
        )
        if response.ok:
            return response.json()
        st.error(f"Ошибка: {response.json().get('detail', 'Неизвестная ошибка')}")
        return None
    except Exception as e:
        st.error(f"Ошибка: {e}")
        return None


def bulk_disable_preview_file(file) -> Optional[dict]:
    try:
        response = requests.post(
            f"{API_URL}/api/v1/users/bulk-disable-preview",
            files={"file": file},
            cookies=get_cookies()
        )
        if response.ok:
            return response.json()
        st.error(f"Ошибка: {response.json().get('detail', 'Неизвестная ошибка')}")
        return None
    except Exception as e:
        st.error(f"Ошибка: {e}")
        return None


def bulk_disable_list(usernames: List[str]) -> Optional[dict]:
    try:
        response = requests.post(
            f"{API_URL}/api/v1/users/bulk-disable",
            json=usernames,
            cookies=get_cookies()
        )
        if response.ok:
            return response.json()
        st.error(f"Ошибка: {response.json().get('detail', 'Неизвестная ошибка')}")
        return None
    except Exception as e:
        st.error(f"Ошибка: {e}")
        return None


def bulk_enable_list(usernames: List[str]) -> Optional[dict]:
    try:
        response = requests.post(
            f"{API_URL}/api/v1/users/bulk-enable",
            json=usernames,
            cookies=get_cookies()
        )
        if response.ok:
            return response.json()
        st.error(f"Ошибка: {response.json().get('detail', 'Неизвестная ошибка')}")
        return None
    except Exception as e:
        st.error(f"Ошибка: {e}")
        return None


def get_groups() -> List[str]:
    try:
        response = requests.get(
            f"{API_URL}/api/v1/groups",
            cookies=get_cookies()
        )
        if response.ok:
            return response.json().get("groups", [])
        return []
    except Exception:
        return []


def create_user(first_name, last_name, email, phone="", title="", groups="") -> Optional[dict]:
    try:
        response = requests.post(
            f"{API_URL}/api/v1/users/create-form",
            data={
                "first_name": first_name,
                "last_name": last_name,
                "email": email,
                "phone": phone,
                "title": title,
                "groups": groups
            },
            cookies=get_cookies()
        )
        if response.ok:
            return response.json()
        st.error(f"Ошибка: {response.json().get('detail', 'Неизвестная ошибка')}")
        return None
    except Exception as e:
        st.error(f"Ошибка: {e}")
        return None


def search_users(q: str) -> Optional[dict]:
    try:
        response = requests.get(
            f"{API_URL}/api/v1/users/search",
            params={"q": q},
            cookies=get_cookies()
        )
        if response.ok:
            return response.json()
        st.error(f"Ошибка: {response.json().get('detail', 'Неизвестная ошибка')}")
        return None
    except Exception as e:
        st.error(f"Ошибка: {e}")
        return None


def reset_user_password(username: str) -> Optional[dict]:
    try:
        response = requests.post(
            f"{API_URL}/api/v1/users/{username}/reset-password",
            cookies=get_cookies()
        )
        if response.ok:
            return response.json()
        st.error(f"Ошибка: {response.json().get('detail', 'Неизвестная ошибка')}")
        return None
    except Exception as e:
        st.error(f"Ошибка: {e}")
        return None


def toggle_user_state(username: str, disable: bool) -> Optional[dict]:
    action = "disable" if disable else "enable"
    try:
        response = requests.post(
            f"{API_URL}/api/v1/users/{username}/{action}",
            cookies=get_cookies()
        )
        if response.ok:
            return response.json()
        st.error(f"Ошибка: {response.json().get('detail', 'Неизвестная ошибка')}")
        return None
    except Exception as e:
        st.error(f"Ошибка: {e}")
        return None


def bulk_create_from_excel(file) -> Optional[dict]:
    try:
        response = requests.post(
            f"{API_URL}/api/v1/users/bulk-create-from-excel",
            files={"file": file},
            cookies=get_cookies()
        )
        if response.ok:
            return response.json()
        st.error(f"Ошибка: {response.json().get('detail', 'Неизвестная ошибка')}")
        return None
    except Exception as e:
        st.error(f"Ошибка: {e}")
        return None


# === HELPERS ===

def parse_textarea(text: str) -> List[str]:
    return [line.strip() for line in text.strip().split('\n') if line.strip()]


def build_reset_excel(result: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сброс паролей"

    headers = ["Username", "Email", "Yopass ссылка", "Пароль", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for user in result["success"]:
        yopass_link = user.get("yopass_link", "")
        ws.cell(row=row, column=1, value=user["username"])
        ws.cell(row=row, column=2, value=user.get("email", ""))
        ws.cell(row=row, column=3, value=yopass_link)
        ws.cell(row=row, column=4, value=user.get("password", "") if not yopass_link else "")
        ws.cell(row=row, column=5, value="Успешно")
        row += 1

    for user in result["failed"]:
        ws.cell(row=row, column=1, value=user["identifier"])
        ws.cell(row=row, column=2, value="")
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=5, value=f"Ошибка: {user['error']}")
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 85
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 20

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def show_reset_results(result: dict):
    total = len(result["success"]) + len(result["failed"])
    col1, col2, col3 = st.columns(3)
    col1.metric("Всего", total)
    col2.metric("Успешно", len(result["success"]))
    col3.metric("Ошибок", len(result["failed"]))

    if result["success"]:
        st.download_button(
            label="Скачать Excel с результатами",
            data=build_reset_excel(result),
            file_name="password_reset_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    st.markdown("---")
    for idx, user in enumerate(result["success"]):
        email_str = f" · {user['email']}" if user.get("email") else ""
        is_blocked = user.get("status") == "disabled"
        status_icon = "🔴" if is_blocked else "🟢"
        yopass_link = user.get("yopass_link", "")
        c1, c2 = st.columns([1, 2])
        with c1:
            st.success(f"✅ {status_icon} **{user['username']}**{email_str}")
            if is_blocked:
                st.warning("⚠️ Пользователь заблокирован — разблокируйте после сброса")
            st.caption("🔑 Требует смены пароля при входе")
        with c2:
            if yopass_link:
                st.text_input(
                    "",
                    value=yopass_link,
                    disabled=True,
                    label_visibility="collapsed",
                    key=f"yp_{idx}_{user['username']}"
                )
            else:
                st.code(user.get("password", ""))
                st.caption("Сохраните пароль — он виден только сейчас")
    for fail in result["failed"]:
        st.error(f"❌ **{fail['identifier']}** — {fail['error']}")


# === PAGES ===

def page_passwords():
    text_input = st.text_area(
        "Вставьте username или email (каждый с новой строки)",
        placeholder="ivan.ivanov\npetr@company.com\nmaria.sidorova",
        height=130,
        key="pwd_text_input"
    )

    with st.expander("Или загрузите Excel (колонка A = username или email, строка 1 — заголовок)"):
        st.markdown(f"[Скачать шаблон Excel]({API_URL}/api/v1/templates/templates-excel-worksusers)")
        excel_file = st.file_uploader("Excel файл", type=["xlsx"], key="pwd_excel_file")

    if st.button("Проверить список", use_container_width=True, key="pwd_preview_btn"):
        identifiers = parse_textarea(text_input)
        preview = None

        if excel_file:
            with st.spinner("Проверяем..."):
                preview = bulk_disable_preview_file(excel_file)
        elif identifiers:
            with st.spinner("Проверяем..."):
                preview = bulk_disable_preview_list(identifiers)
        else:
            st.warning("Введите пользователей или загрузите файл")

        if preview:
            st.session_state.pwd_preview = preview
            st.session_state.pwd_results = None

    # Шаг 1: показываем превью
    if st.session_state.pwd_preview:
        preview = st.session_state.pwd_preview
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        col1.metric("Всего в списке", preview["total"])
        col2.metric("Найдено", preview["found_count"])
        col3.metric("Не найдено", preview["not_found_count"])
        st.markdown("---")

        for u in preview["found"]:
            st.success(f"✅ `{u['identifier']}` → **{u['username']}**")
        for u in preview["not_found"]:
            st.error(f"❌ `{u['identifier']}` — {u['error']}")

        if preview["found_count"] > 0:
            st.markdown("---")
            st.warning(f"Пароли будут сброшены для **{preview['found_count']}** пользователей")
            skip_yopass = st.checkbox(
                "Показать пароли напрямую (без Yopass)",
                value=st.session_state.show_passwords_mode,
                key="pwd_skip_yopass",
                help="Yopass не будет использован — пароль будет показан в открытом виде"
            )
            st.session_state.show_passwords_mode = skip_yopass
            if st.button("Сбросить пароли", use_container_width=True, key="pwd_reset_btn", type="primary"):
                identifiers = [u["identifier"] for u in preview["found"]]
                with st.spinner("Сбрасываем пароли..."):
                    if skip_yopass:
                        result = bulk_reset_plain(identifiers)
                    else:
                        result = bulk_reset_with_yopass(identifiers)
                if result:
                    st.session_state.pwd_results = result
                    st.session_state.pwd_preview = None
                    st.rerun()

    # Шаг 2: показываем результаты
    if st.session_state.pwd_results:
        st.markdown("---")
        show_reset_results(st.session_state.pwd_results)


def page_state_management():
    mode = st.radio("", ["Заблокировать", "Разблокировать"], horizontal=True, key="state_mode")
    is_disable = mode == "Заблокировать"
    st.markdown("---")

    # Если режим сменился — сбрасываем preview/results предыдущего режима
    if st.session_state.state_mode_used and st.session_state.state_mode_used != mode:
        st.session_state.state_preview = None
        st.session_state.state_results = None

    text_input = st.text_area(
        "Вставьте username или email (каждый с новой строки)",
        placeholder="ivan.ivanov\npetr@company.com\nmaria.sidorova",
        height=130,
        key="state_text_input"
    )

    with st.expander("Или загрузите Excel (колонка A = username или email, строка 1 — заголовок)"):
        st.markdown(f"[Скачать шаблон Excel]({API_URL}/api/v1/templates/templates-excel-worksusers)")
        excel_file = st.file_uploader("Excel файл", type=["xlsx"], key="state_excel_file")

    if st.button("Проверить список", use_container_width=True, key="state_preview_btn"):
        identifiers = parse_textarea(text_input)
        result = None

        if excel_file:
            with st.spinner("Проверяем..."):
                result = bulk_disable_preview_file(excel_file)
        elif identifiers:
            with st.spinner("Проверяем..."):
                result = bulk_disable_preview_list(identifiers)
        else:
            st.warning("Введите пользователей или загрузите файл")

        if result:
            st.session_state.state_preview = result
            st.session_state.state_results = None
            st.session_state.state_mode_used = mode

    # Показываем результат проверки
    if st.session_state.state_preview:
        preview = st.session_state.state_preview
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        col1.metric("Всего в списке", preview["total"])
        col2.metric("Найдено", preview["found_count"])
        col3.metric("Не найдено", preview["not_found_count"])
        st.markdown("---")

        for u in preview["found"]:
            st.success(f"✅ `{u['identifier']}` → **{u['username']}**")
        for u in preview["not_found"]:
            st.error(f"❌ `{u['identifier']}` — {u['error']}")

        if preview["found_count"] > 0:
            st.markdown("---")
            action_label = "заблокировано" if is_disable else "разблокировано"
            btn_label = "Заблокировать" if is_disable else "Разблокировать"
            st.warning(f"Будет {action_label} **{preview['found_count']}** пользователей")
            if st.button(btn_label, use_container_width=True, key="state_exec_btn", type="primary"):
                usernames = [u["username"] for u in preview["found"]]
                spinner_text = "Блокируем..." if is_disable else "Разблокируем..."
                with st.spinner(spinner_text):
                    result = bulk_disable_list(usernames) if is_disable else bulk_enable_list(usernames)
                if result:
                    st.session_state.state_results = result
                    st.session_state.state_mode_used_exec = mode
                    st.session_state.state_preview = None
                    st.rerun()

    # Показываем итоговый результат
    if st.session_state.state_results:
        result = st.session_state.state_results
        exec_mode = st.session_state.state_mode_used_exec or mode
        action_done = "Заблокировано" if exec_mode == "Заблокировать" else "Разблокировано"
        action_verb = "заблокирован" if exec_mode == "Заблокировать" else "разблокирован"
        already_list = result.get("already", [])
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        col1.metric(action_done, len(result["success"]))
        col2.metric("Уже в нужном состоянии", len(already_list))
        col3.metric("Ошибок", len(result["failed"]))
        st.markdown("---")
        for u in result["success"]:
            st.success(f"✅ **{u['username']}** — {action_verb}")
        for u in already_list:
            st.warning(f"⚠️ **{u['username']}** — {u['note']}")
        for u in result["failed"]:
            st.error(f"❌ **{u['identifier']}** — {u['error']}")


def page_create_single():
    # Загружаем группы один раз и кешируем
    if st.session_state.groups_list is None:
        with st.spinner("Загружаем список групп..."):
            st.session_state.groups_list = get_groups()

    col_refresh, _ = st.columns([1, 4])
    with col_refresh:
        if st.button("↻ Обновить группы", key="refresh_groups", use_container_width=True):
            st.session_state.groups_list = get_groups()
            st.rerun()

    with st.form("create_single_form"):
        col1, col2 = st.columns(2)
        with col1:
            first_name = st.text_input("Имя *", placeholder="Иван")
            email = st.text_input("Email *", placeholder="ivan@company.com")
            title = st.text_input("Должность", placeholder="Менеджер")
        with col2:
            last_name = st.text_input("Фамилия *", placeholder="Иванов")
            phone = st.text_input("Телефон", placeholder="+7 999 123 45 67")
            selected_groups = st.multiselect(
                "Группы",
                options=st.session_state.groups_list or [],
                placeholder="Выберите группы из списка..."
            )

        submitted = st.form_submit_button("Создать пользователя", use_container_width=True)

        if submitted:
            if not first_name or not last_name or not email:
                st.error("Заполните обязательные поля: Имя, Фамилия, Email")
            else:
                groups_str = ",".join(selected_groups)
                result = create_user(first_name, last_name, email, phone, title, groups_str)
                if result:
                    st.success(f"Пользователь **{result['username']}** создан!")
                    if result.get("yopass_link"):
                        st.text_input(
                            "Yopass ссылка:",
                            value=result["yopass_link"],
                            disabled=True,
                            label_visibility="collapsed",
                            key="create_single_yopass"
                        )
                        st.info("Ссылка одноразовая, действует 7 дней")
                    if result.get("groups"):
                        if result["groups"].get("added"):
                            st.success(f"Добавлен в группы: {', '.join(result['groups']['added'])}")
                        if result["groups"].get("failed"):
                            st.warning(f"Не удалось добавить в группы: {result['groups']['failed']}")


def page_create_excel():
    st.markdown(f"**Шаг 1:** [Скачать шаблон Excel]({API_URL}/api/v1/templates/templates-excel)")
    st.markdown("**Шаг 2:** Загрузите заполненный файл")

    uploaded_file = st.file_uploader("Excel файл", type=["xlsx"], key="create_excel_file")

    if uploaded_file is not None:
        if st.button("Создать пользователей", use_container_width=True, key="create_excel_btn"):
            with st.spinner("Создаём пользователей..."):
                result = bulk_create_from_excel(uploaded_file)

            if result:
                col1, col2 = st.columns(2)
                col1.metric("Создано", len(result["success"]))
                col2.metric("Ошибок", len(result["failed"]))
                st.markdown("---")
                for idx, user in enumerate(result["success"]):
                    c1, c2 = st.columns([1, 2])
                    with c1:
                        st.success(f"✅ **{user['username']}**\n{user['email']}")
                    with c2:
                        st.text_input(
                            "",
                            value=user["yopass_link"],
                            disabled=True,
                            label_visibility="collapsed",
                            key=f"create_yopass_{idx}"
                        )
                for fail in result["failed"]:
                    st.error(f"❌ Строка {fail['row']}: {fail['error']}")


def page_reports():
    st.markdown("### Экспорт пользователей и групп в CSV")

    if st.button("Сформировать CSV", use_container_width=True, key="reports_btn"):
        with st.spinner("Генерируем отчёт..."):
            try:
                response = requests.get(
                    f"{API_URL}/api/v1/report/full-usersgroups-info",
                    cookies=get_cookies(),
                    timeout=30
                )
                if response.ok:
                    st.session_state.csv_data = response.content
                    st.success("Отчёт готов!")
                else:
                    st.error(f"Ошибка: {response.status_code}")
            except Exception as e:
                st.error(f"Ошибка: {e}")

    if st.session_state.csv_data:
        st.download_button(
            label="Скачать CSV",
            data=st.session_state.csv_data,
            file_name="users_groups_report.csv",
            mime="text/csv",
            use_container_width=True
        )
        st.caption("Файл содержит: username, email, список групп")


def page_search_users():
    with st.form("search_form"):
        col1, col2 = st.columns([4, 1])
        with col1:
            q = st.text_input(
                "",
                placeholder="Введите имя, фамилию, username или email...",
                label_visibility="collapsed"
            )
        with col2:
            submitted = st.form_submit_button("Найти", use_container_width=True)

    if submitted:
        if not q.strip():
            st.warning("Введите поисковый запрос")
        else:
            with st.spinner("Ищем..."):
                result = search_users(q.strip())
            if result is not None:
                st.session_state.search_results = result
                st.session_state.search_action_result = None

    if st.session_state.search_results is not None:
        result = st.session_state.search_results
        st.caption(f"Найдено: {result['count']} · Кликните на строку чтобы выбрать пользователя")

        if result["count"] == 0:
            st.info("Ничего не найдено")
            return

        rows = []
        for u in result["users"]:
            status_str = "🟢 Активен" if u["status"] == "active" else "🔴 Заблокирован"
            groups_str = ", ".join(u["groups"]) if u["groups"] else "—"
            rows.append({
                "Статус": status_str,
                "Username": u["username"],
                "ФИО": u["full_name"],
                "Email": u["email"],
                "Должность": u["title"] or "—",
                "Группы": groups_str,
            })

        event = st.dataframe(
            rows,
            use_container_width=True,
            hide_index=True,
            selection_mode="single-row",
            on_select="rerun",
            key="search_table"
        )

        selected_rows = event.selection.rows
        if selected_rows:
            selected_idx = selected_rows[0]
            user = result["users"][selected_idx]
            is_disabled = user["status"] == "disabled"

            st.markdown("---")
            col_info, col_r, col_t = st.columns([3, 1, 1])
            with col_info:
                name = user["full_name"] or user["username"]
                st.markdown(f"**{name}** · `{user['username']}` · {user['email']}")
                skip_yopass = st.checkbox(
                    "Показать пароль напрямую (без Yopass)",
                    value=st.session_state.show_passwords_mode,
                    key="search_skip_yopass",
                    help="Пароль будет показан в открытом виде, без создания Yopass ссылки"
                )
                st.session_state.show_passwords_mode = skip_yopass
            with col_r:
                if st.button("🔑 Сбросить пароль", use_container_width=True, key="search_reset_btn"):
                    with st.spinner("Сбрасываем..."):
                        if st.session_state.show_passwords_mode:
                            bulk_res = bulk_reset_plain([user["username"]])
                            res = bulk_res["success"][0] if bulk_res and bulk_res.get("success") else None
                            if res:
                                res = {"username": res["username"], "password": res["password"], "yopass_link": ""}
                        else:
                            res = reset_user_password(user["username"])
                    if res:
                        st.session_state.search_action_result = {"type": "reset", "data": res}
                        st.rerun()
            with col_t:
                if is_disabled:
                    if st.button("✅ Разблокировать", use_container_width=True, key="search_toggle_btn"):
                        with st.spinner("Разблокируем..."):
                            res = toggle_user_state(user["username"], disable=False)
                        if res:
                            st.session_state.search_results["users"][selected_idx]["status"] = "active"
                            st.session_state.search_action_result = {"type": "unblocked", "username": user["username"]}
                            st.rerun()
                else:
                    if st.button("🔒 Заблокировать", use_container_width=True, key="search_toggle_btn"):
                        with st.spinner("Блокируем..."):
                            res = toggle_user_state(user["username"], disable=True)
                        if res:
                            st.session_state.search_results["users"][selected_idx]["status"] = "disabled"
                            st.session_state.search_action_result = {"type": "blocked", "username": user["username"]}
                            st.rerun()

        if st.session_state.search_action_result:
            action_res = st.session_state.search_action_result
            st.markdown("---")
            if action_res["type"] == "reset":
                data = action_res["data"]
                st.success(f"✅ Пароль сброшен для **{data['username']}**")
                if data.get("yopass_link"):
                    st.text_input(
                        "",
                        value=data["yopass_link"],
                        disabled=True,
                        label_visibility="collapsed",
                        key="search_yopass"
                    )
                    st.caption("Ссылка одноразовая, действует 7 дней")
                else:
                    st.code(data.get("password", ""))
                    st.caption("Сохраните пароль — он виден только сейчас")
            elif action_res["type"] == "blocked":
                st.success(f"✅ **{action_res['username']}** заблокирован")
            elif action_res["type"] == "unblocked":
                st.success(f"✅ **{action_res['username']}** разблокирован")


# === APP CONFIG ===
st.set_page_config(page_title="FreeIPA Portal", page_icon="", layout="wide")
st.title("FreeIPA Portal")

# === LOGIN ===
if not st.session_state.logged_in:
    st.markdown("---")
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.subheader("Вход в систему")
        with st.form("login_form"):
            username = st.text_input("Username", key="login_username")
            password = st.text_input("Password", type="password", key="login_password")
            submitted = st.form_submit_button("Войти", use_container_width=True)
            if submitted:
                if login(username, password):
                    st.success("Вы авторизованы")
                    st.rerun()

# === MAIN ===
else:
    col1, col2 = st.columns([3, 1])
    with col1:
        st.markdown(f"**Вы:** {st.session_state.username}")
    with col2:
        if st.button("Выйти", use_container_width=True, key="logout_btn"):
            logout()
            st.rerun()

    st.markdown("---")

    tab1, tab2, tab3, tab4 = st.tabs(["Пароли", "Управление состояниями", "Пользователи", "Аналитика"])

    with tab1:
        page_passwords()

    with tab2:
        page_state_management()

    with tab3:
        mode = st.radio("", ["Создать одного", "Создать из Excel", "Поиск"], horizontal=True, key="user_mode")
        st.markdown("---")
        if mode == "Создать одного":
            page_create_single()
        elif mode == "Создать из Excel":
            page_create_excel()
        else:
            page_search_users()

    with tab4:
        page_reports()

st.markdown("---")
st.markdown("*FreeIPA Portal — Управление пользователями*")
